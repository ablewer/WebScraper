import openpyxl
from openpyxl.utils import get_column_letter

x = [['test', 'object'], ['second', 'joined']]

wb = openpyxl.load_workbook('Excel_Data.xlsx')

sheets = wb.sheetnames
for item in sheets:
    if item != 'Sheet':
        current_sheet = wb[item]
        wb.remove(current_sheet)

active_sheet = wb.active
print(active_sheet)

sheets = wb.sheetnames
print(sheets)

for index in range(len(x)):
    string = 'Sheet' + str(index)
    wb.create_sheet(string)

sheets = wb.sheetnames
print(sheets)

wb.save('Excel_Data.xlsx')

wb.close()
