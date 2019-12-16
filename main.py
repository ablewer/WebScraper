'''
This program's function is to pull data about jobs and careers specified by the user and organizes the data into an
excel file.

Using the googlesearch module made by Mario Vilas at https://breakingcode.wordpress.com/
'''

import sys, requests, bs4, os, re, pprint, multiprocessing, openpyxl, docx
from multiprocessing import freeze_support, Manager
from openpyxl.utils import get_column_letter  # getting the get_colum_letter function
from googlesearch import search
from ResumeTest import get_resume_text


# functions
# this function will take the URL's and find the HTML data, saving each to a Dictionary
def parse_data(id, q, d):
    while True:
        item = q.get()
        if item is None:
            break

        assert isinstance(item, str), 'Error: Item in func parse_data is not a string'
        assert isinstance(id, int), 'Error: id in func parse_data is not an int'

        print('Scraping url: ' + item + ' Thread: ' + (str(id) + ' open.'))

        new_res = requests.get(item)  # new request pull of this url

        html_soup = bs4.BeautifulSoup(new_res.text, 'html.parser')  # parse the html of the url into an object

        new_elems = html_soup.select('title')  # find the element named title

        title = ''  # create an empty string for use later

        if new_elems:  # if the object is not None
            for j in range(len(new_elems)):  # for each index value in range of the amount of items found (should be 1)
                title = new_elems[j].text  # title is now equal to the text of this object

        # find the technology related key words
        mo = technology_regex.findall(new_res.text.lower())  # search the html for all technology key words
        tech_count = 0  # tech keyword search count set to an initial number of zero
        if mo:  # if mo is not None
            for j in mo:  # for each object in mo
                tech_count += 1  # add one to the tech count search

        # find the careers related key words
        mo = careers_regex.findall(new_res.text.lower())  # search the html for career key words
        career_count = 0  # career_count set to zero
        if mo:  # if mo is not None
            for j in mo:  # for each object in mo
                career_count += 1  # add one to the career count search

        # find the email related formated strings
        mo = email_regex.findall(new_res.text.lower())  # search the html for emails
        email_list = []  # create an empty list to add emails to
        if mo:  # if mo is not None
            for j in mo:  # for each object in mo
                email_list.append(j[0])  # add that item to the email list

        # find the computer related key words
        computer_count = 0  # computer_count set to zero
        mo = computer_regex.findall(new_res.text.lower())  # search the html for computer keywords
        if mo:  # if the search is not None
            for j in mo:  # for each j in mo
                computer_count += 1  # add one to the computer count search

        # create the reference dictionary with the data for each item
        d[title] = {'Tech': tech_count, 'Career': career_count, 'Email List': email_list,
                    'Computer': computer_count, 'url': item}

        print("Scraping done. Thread: " + (str(id) + " closed"))


# regex to search for technology keywords
technology_regex = re.compile(r'''
    (tech|technology)
''', re.VERBOSE)

# regex to search for computer keywords
computer_regex = re.compile(r'''
    (computer|computer science)
''', re.VERBOSE)

# regex to search for career keywords
careers_regex = re.compile(r'''
    (careers|career|jobs|job|professional|hiring|student)
''', re.VERBOSE)

# regex to search for emails
email_regex = re.compile(r'''
    ([a-zA-Z0-9.]+@\w+\.(com|ord|edu|net))
''', re.VERBOSE)

# start consumers
if __name__ == '__main__':

    # read in the resume
    x = []
    techSummary = {}
    fileNames = []
    for root, dirs, files in os.walk("."):
        for filename in files:
            if filename.endswith("Resume.docx"):
                fileNames.append(filename)

    for resume in fileNames:
        doc = docx.Document(resume)
        for para_index in range(len(doc.paragraphs)):
            assert isinstance(para_index, int), 'Error: para_index is not an int'
            # search a few different types of resumes
            if doc.paragraphs[para_index].text.lower() == 'technology summary' or \
                    doc.paragraphs[para_index].text.lower() =='skills' or \
                    doc.paragraphs[para_index].text.lower() =='technical proficiency' or \
                    doc.paragraphs[para_index].text.lower() =='technical summary':
                var_string = ''
                for string in doc.paragraphs[para_index + 1].runs:
                    var_string += string.text
                techSummary[resume] = var_string

    for value in techSummary.values():
        assert isinstance(value, str), 'Error: techSummary.values() does not return a string'
        data = value.split()
        x.append(data)

    # checks for command arguements
    if sys.argv.__len__() > 1:
        if isinstance(sys.argv[1], int):  # if 1 element is an integer
            numProcesses = sys.argv[1]
            string = '' + ' '.join(sys.argv[2:])  # use the 3rd element on
        else:  # if the 1 element is not integer
            string = '' + ' '.join(sys.argv[1:])  # use the 2nd element on
    else:
        string = 'jobs'  # start of searching string
        print('Use: ')

    wb = openpyxl.load_workbook('Excel_Data.xlsx')  # load the excel file
    sheets = wb.sheetnames  # load the names of all the sheets
    for item in sheets:  # for every item in the list of sheets
        if item != 'Sheet':  # if the name is not sheets
            current_sheet = wb[item]  # get the sheet object
            wb.remove(current_sheet)  # remove it from the work book
    for item in techSummary.keys():  # for index in range of the length of lists of lists
        string = str(item)  # create a string with sheet and the number of the index
        wb.create_sheet(string)  # create a new sheet object in the work book with this name
    wb.save('Excel_Data.xlsx')  # save this to the file
    wb.close()  # close the file

    # setting up variables before function
    stop_num = 5  # int for number of items google searches for, the default number
    resume_name_list = list(techSummary.keys())  # create a list for the names of the resume from the dictionary

    for list_item in x:  # for every list in the list of lists

        assert isinstance(list_item, list), 'Error: list_item is not a list'

        # variables
        url_que = multiprocessing.Queue()  # initialize the que
        processes = []  # array for the process
        numProcesses = multiprocessing.cpu_count()  # number of process that will be used

        search_string = ''  # empty search string
        for item in list_item:  # for every item inside the list
            search_string += str(item) + ' OR '  # add every item to the empty string plus an OR
        search_string += 'AND software developer jobs'  # end the search string with and software developer jobs
        stop_num = round(35/len(x))  # set the stop number equal to that of 35 / length of the list of lists

        start_num = 1  # starting number for percentage output
        for data in search(search_string, stop=stop_num):  # for each piece of data in the search that stops at 20
            url_que.put(data)  # place the data into a que
            print(str(round((start_num / stop_num) * 100, 0)) + '% done searching for urls')
            start_num += 1

        manager = Manager()
        d = manager.dict()

        for processId in range(numProcesses):
            processes.append(multiprocessing.Process(target=parse_data, args=(processId, url_que, d)))
            processes[processId].start()
            url_que.put(None)

        for p in processes:
            p.join()

        pprint.pprint(d)  # pprint the dictionary

        excel_file = openpyxl.load_workbook('Excel_Data.xlsx')  # load the excel file

        sheet_string = resume_name_list[x.index(list_item)]  # get the string of the resume name
        sheet = excel_file[sheet_string]  # get the sheet wanted

        # create a list for the titles of the columns
        title_list = ['Title', 'URL', 'Career', 'Computer', 'Tech', 'Email']

        col_num = 1  # set column number to 1 or A

        for value in title_list:  # for each value in the title list
            cell = str(get_column_letter(col_num) + str(1))  # set the cell value equal to the current column at row 1
            sheet[cell] = value  # change the value of that cell to the value in the title list
            col_num += 1  # add one to the row number to change the column

        row_num = 2  # set the row number to 2 or B
        for value in d:  # for each value in the reference dictionary
            cell = str(
                get_column_letter(1) + str(row_num))  # get the cell equal to the first column current row fo iteration
            sheet[cell] = value  # change that cell to that value

            cell = str(get_column_letter(6) + str(row_num))  # cell at second column current row
            message = ''  # empty string
            for item in d[value]['Email List']:  # for each item in the email list
                message += str(item + '\n')  # add the email to the message plus a new line character
            sheet[cell] = message  # set the cell equal to message

            # set the 3rd column current row equal to career count
            cell = str(get_column_letter(3) + str(row_num))
            sheet[cell] = d[value]['Career']

            # set the 4th column current row equal to Computer count
            cell = str(get_column_letter(4) + str(row_num))
            sheet[cell] = d[value]['Computer']

            # set the 5th column current row equal to Tech count
            cell = str(get_column_letter(5) + str(row_num))
            sheet[cell] = d[value]['Tech']

            # set the 6th column current row equal to the url
            cell = str(get_column_letter(2) + str(row_num))
            sheet[cell] = d[value]['url']

            # increase the column number
            row_num += 1

        excel_file.save('Excel_Data.xlsx')  # save the excel file

    emailDict = {}
    sheetDict = {}
    for sheet in sheets:
        print(sheet)
        for row in range(2, wb[sheet].max_row + 1):
            title = wb[sheet]['A' + str(row)].value
            email = wb[sheet]['F' + str(row)].value
            sheetDict[title] = [email]
        emailDict[sheet] = sheetDict

    pprint.pprint(emailDict)


    excel_file.close()  # close the excel file

    print('Data saved to Excel_Data.xlsx Successfully')
