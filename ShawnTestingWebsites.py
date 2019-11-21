'''
This program's function is to pull data about jobs and careers specified by the user and organizes the data into an
excel file.
'''

import sys  # for use in the sys.argv
import requests  # for use in getting the web page
import bs4  # for use of parsing the html file
import os  # for use of system
import re  # for use of searching
import pprint  # for use of pprinting
import openpyxl  # for use of excel files
from openpyxl.utils import get_column_letter  # getting the get_colum_letter function

# regex to search for technology keywords
technology_regex = re.compile(r'''
    (tech|technology)
''', re.VERBOSE)

# regex to search for computer keywords
computer_regex = re.compile(r'''
    (computer|computer science)
''', re.VERBOSE)

# regex to search for career keywords
careers_regex = re.compile(r'''
    (careers|career|jobs|job|professional|hiring|student)
''', re.VERBOSE)

# regex to search for emails
email_regex = re.compile(r'''
    ([a-zA-Z0-9]+@[a-zA-Z0-9.]+)
''', re.VERBOSE)

string = 'https://www.google.com/search?q=cars'  # start of searching string

print('using: ' + string)  # print out to the user what exact search it is doing

res = requests.get(string)  # requests module pulls the html from the url

soup = bs4.BeautifulSoup(res.text, "html.parser")  # create a soup obj that parses the html from the request

p_elems = soup.select('div[class="kCrYT"] > a')  # select the urls on the google page

rank_dict = {}  # empty dictionary for data to be dumped into

for i in range(len(p_elems)):  # for each index value in the length of the search result urls
    word = p_elems[i].get('href')  # get the url out of the html element
    index_value = word.find('&')  # find the index of the & symbol as to remove the extra data
    url = p_elems[i].get('href')[7:int(index_value)]  # change the url to get the url minus the extra data
    print('\nResult #' + str(i) + ': ' + url)  # output to the user

    new_res = requests.get(url)  # new request pull of this url

    html_soup = bs4.BeautifulSoup(new_res.text, 'html.parser')  # parse the html of the url into an object

    new_elems = html_soup.select('title')  # find the element named title

    title = ''  # create an empty string for use later

    if new_elems:  # if the object is not None
        for j in range(len(new_elems)):  # for each index value j in range of the amount of items found (should be 1)
            title = new_elems[j].text  # title is now equal to the text of this object

    mo = technology_regex.findall(new_res.text.lower())  # search the html for all technology key words

    tech_count = 0  # tech keyword search count set to an initial number of zero

    if mo:  # if mo is not None
        for j in mo:  # for each object in mo
            tech_count += 1  # add one to the tech count search

    mo = careers_regex.findall(new_res.text.lower())  # search the html for career key words

    career_count = 0  # career_count set to zero

    if mo:  # if mo is not None
        for j in mo:  # for each object in mo
            career_count += 1  # add one to the career count search

    mo = email_regex.findall(new_res.text.lower())  # search the html for emails

    email_list = []  # create an empty list to add emails to

    if mo:  # if mo is not None
        for j in mo:  # for each object in mo
            email_list.append(j)  # add that item to the email list

    computer_count = 0  # computer_count set to zero

    mo = computer_regex.findall(new_res.text.lower())  # search the html for computer keywords

    if mo:  # if the search is not None
        for j in mo:  # for each j in mo
            computer_count += 1  # add one to the computer count search

    print('Tech Search hit count: ' + str(tech_count))  # output to user (can be deleted)
    print('Career Search hit count: ' + str(career_count))  # output to user (can be deleted)
    print('Computer Search hit Count: ' + str(computer_count))  # output to user (can be deleted)

    print('Email List:')   # output to user (can be deleted)
    for email in email_list:   # output to user (can be deleted)
        print('\t' + email)

    # create the reference dictionary with the data for each item
    rank_dict[title] = {'Tech': tech_count, 'Career': career_count, 'Email List': email_list,
                        'Computer': computer_count, 'url': url}

pprint.pprint(rank_dict)  # pprint the dictionary

excel_file = openpyxl.Workbook()  # create an empty dictionary

sheet = excel_file.active  # get the active sheet

# create a list for the titles of the columns
title_list = ['Title', 'Email', 'Career', 'Computer', 'Tech', 'URL']

iteration = 1  # set iteration to 1

for value in title_list:  # for each value in the title list
    cell = str(get_column_letter(iteration) + str(1))  # set the cell value equal to the current column at row 1
    sheet[cell] = value  # change the value of that cell to the value in the title list
    iteration += 1  # add one to the iteration to change the column

iteration = 2  # set the iteration 2
for value in rank_dict:  # for each value in the reference dictionary
    cell = str(get_column_letter(1) + str(iteration))  # get the cell equal to the first column current row fo iteration
    sheet[cell] = value  # change that cell to that value

    cell = str(get_column_letter(2) + str(iteration))  # cell at second column current row
    message = ''  # empty string
    for item in rank_dict[value]['Email List']:  # for each item in the email list
        message += str(item + '\n')  # add the email to the message plus a new line character
    sheet[cell] = message  # set the cell equal to message

    # set the 3rd column current row equal to career count
    cell = str(get_column_letter(3) + str(iteration))
    sheet[cell] = rank_dict[value]['Career']

    # set the 4th column current row equal to Computer count
    cell = str(get_column_letter(4) + str(iteration))
    sheet[cell] = rank_dict[value]['Computer']

    # set the 5th column current row equal to Tech count
    cell = str(get_column_letter(5) + str(iteration))
    sheet[cell] = rank_dict[value]['Tech']

    # set the 6th column current row equal to the url
    cell = str(get_column_letter(6) + str(iteration))
    sheet[cell] = rank_dict[value]['url']

    # make the iteration move one
    iteration += 1

excel_file.save('ShawnTestData.xlsx')  # save the excel file

excel_file.close()  # close the excel file
